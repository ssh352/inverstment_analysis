#!/usr/bin/env python
# coding:utf-8

"""
function: 
@author: zkang kai
@contact: 474918208@qq.com
"""

import wx
import wx.richtext as rt
import requests
import pandas as pd
from bs4 import BeautifulSoup
import tushare as ts
import openpyxl as oxl
import os
from collections import OrderedDict
import datetime
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np



def load_config_data(sh_code):
    '''
    根据股票代码返回数据
    '''
    config_json = sh_code + '.json'
    if os.path.exists(config_json):
        with open(config_json,'r',encoding='utf-8') as f:
            config_dict = json.load(f)
    else:
        config_dict = {}
        config_dict['name_dict'] = {}
        first_all_data = pd.DataFrame()
        config_dict['first_all_data'] = first_all_data.to_json()
        second_all_data = pd.DataFrame()
        config_dict['second_all_data'] = second_all_data.to_json()
        contents.AppendText("当前未拥有任何数据\n")

    return config_dict

def update(event):
    '''
    更新数据
    '''

    symbol_number = len(sh_code_list)

    begin_date = ui_begin_date.GetValue()
    end_date = ui_end_date.GetValue()
    #  stk_code = symbol_text.GetValue()

    if len(begin_date) != 10 and len(end_date) != 10:
        contents.AppendText("日期格式不对,请参照:20180101填写\n")
        return None

    browser = webdriver.Chrome()
    browser.get(url)

    contents.AppendText('开始下载数据...........\n')

    for i in range(symbol_number):
        sh_code = sh_code_list[i]
        hk_code = hk_code_list[i]

        config_dict = load_config_data(sh_code)

        name_dict = config_dict.get('name_dict',{})
        first_all_data = config_dict['first_all_data']
        first_all_data = pd.read_json(first_all_data)
        second_all_data = config_dict['second_all_data']
        second_all_data = pd.read_json(second_all_data)

        already_begin_date = config_dict.get('already_begin_date',begin_date)
        already_end_date = config_dict.get('already_end_date',end_date)

        temp_trade_cal = global_trade_cal[global_trade_cal['calendarDate'] >= begin_date]
        temp_trade_cal = temp_trade_cal[temp_trade_cal['calendarDate'] <= end_date]
        temp_trade_cal = temp_trade_cal[temp_trade_cal['isOpen'] == 1]
        temp_trade_cal = temp_trade_cal['calendarDate']

        for temp_time in temp_trade_cal:

            if temp_time in config_dict:
                continue

            select = Select(browser.find_element_by_name('ddlShareholdingMonth'))
            select.select_by_value(temp_time[5:7])

            select = Select(browser.find_element_by_name('ddlShareholdingDay'))
            select.select_by_value(temp_time[8:10])

            select = Select(browser.find_element_by_name('ddlShareholdingYear'))
            select.select_by_value(temp_time[0:4])

            element = browser.find_element_by_id("txtStockCode")
            #  element.send_keys("91888")
            element.send_keys(hk_code)

            browser.find_element_by_id("btnSearch").click()

            html_txt = browser.page_source

            soup = BeautifulSoup(html_txt,'html.parser')

            first_part = get_first_part(soup,temp_time)
            second_part = get_second_part(soup,temp_time)
            first_all_data = first_part.append(first_all_data)
            second_all_data = second_part.append(second_all_data)

            config_dict[temp_time] = [first_part.to_json(),second_part.to_json()]
            browser.back()


        first_all_data = first_all_data.sort_index(ascending=False)
        second_all_data = second_all_data.sort_index(ascending=False)
        second_all_data = second_all_data.fillna(0)

        already_begin_date = first_all_data.index[-1].date().isoformat()
        already_end_date = first_all_data.index[0].date().isoformat()
        #  close_data = ts.get_k_data('601888',start=already_begin_date,end=already_end_date)
        close_data = ts.get_k_data(sh_code,start=already_begin_date,end=already_end_date)

        config_dict['name_dict'] = name_dict
        config_dict['close_data'] = close_data.to_json()
        config_dict['already_begin_date'] = already_begin_date
        config_dict['already_end_date'] = already_end_date
        config_dict['first_all_data'] = first_all_data.to_json()
        config_dict['second_all_data'] = second_all_data.to_json()

        with open(sh_code+'.json','w',encoding='utf-8') as f:
            json.dump(config_dict,f)

        contents.AppendText(sh_code+':更新数据完成!\n')

    browser.close()
    contents.AppendText('全部更新完成!\n')


def get_first_part(soup,temp_time):

    temp_html = soup.find("table",{"id":"Table5"})
    temp_html = temp_html.find_all('td',{"class":"arial12black"})

    first_dict = {}
    temp_string = temp_html[1].get_text().strip()
    temp_string = datetime.datetime.strptime(temp_string,"%d/%m/%Y")
    temp_string = temp_string.strftime("%Y-%m-%d")
    #  first_dict['last_date'] = temp_string
    first_dict['stock_code'] = temp_html[3].get_text().strip()
    first_dict['stock_name'] = temp_html[5].get_text().strip()

    temp_html = soup.find("div",{"id":{"pnlResultSummary"}})
    temp_html = temp_html.find_all("span",{"class":"mobilezoom"})

    first_dict = {}
    first_dict['hold_volumn'] = temp_html[0].get_text().strip()
    first_dict['people_number'] = temp_html[1].get_text().strip()
    first_dict['hold_precent'] = temp_html[2].get_text().strip()
    first_dict['all_volumn'] = temp_html[6].get_text().strip()

    return pd.DataFrame(first_dict,index=[pd.Timestamp(temp_time)])

def get_second_part(soup,temp_time):

    name_dict = config_dict.get('name_dict',{})
    temp_html = soup.find("table",{"id":{"participantShareholdingList"}})
    temp_html = temp_html.find_all('tr')

    row_number = len(temp_html)
    col_number = len(temp_html[0].find_all('td'))

    d = {}
    for i in range(col_number):
        d[i] = []

    for i in range(1,row_number):
        temp_tr = temp_html[i]
        temp_td = temp_tr.find_all('td')

        if len(temp_td) !=col_number:
            continue

        for j in range(col_number):
            temp_string = temp_td[j].get_text().strip()
            d[j].append(temp_string)

    last_pd = pd.DataFrame(d)

    last_dict = {}
    for i in range(len(last_pd)):
        last_dict[last_pd.iat[i,0]] = last_pd.iat[i,1]

    name_dict.update(last_dict)

    new_pd = pd.DataFrame(last_pd[3])
    new_pd = new_pd.T
    new_pd.columns = last_pd[0]
    new_pd.index = [pd.Timestamp(temp_time)]

    return new_pd


def load_symbol_data(event):
    '''
    点击按钮加载json当中的数据
    '''
    hk_code_list = []
    sh_code_list = []
    symbol_number = symbol_list.GetNumberOfLines()
    for i in range(symbol_number):
        symbol_code = symbol_list.GetLineText(i)
        if len(symbol_code) == 5:
            hk_code_list.append(symbol_code)
            sh_code_list.append("60"+symbol_code[1:])
        elif len(symbol_code) == 6:
            sh_code_list.append(symbol_code)
            hk_code_list.append("9" + symbol_code[2:])
        else:
            contents.AppendText(symbol_code + ":无法识别股票代码,请填写5位港股或6位上海市场代码\n")



def write_to_excel(event):
    '''
    把数据写入excle中
    '''
    symbol_number = len(sh_code_list)

    for i in range(symbo_numver):

        sh_code = sh_code_list[i]

        config_dict = load_config_data(sh_code)

        name_dict = config_dict.get('name_dict',{})
        first_all_data = config_dict['first_all_data']
        first_all_data = pd.read_json(first_all_data)

        second_all_data = config_dict['second_all_data']
        second_all_data = pd.read_json(second_all_data)

        close_data = config_dict['close_data']
        close_data = pd.read_json(close_data)
        close_data = close_data['close']

        hold_volumn = first_all_data['hold_volumn']
        people_number = first_all_data['people_number']
        hold_precent = first_all_data['hold_precent']
        all_volumn = first_all_data['all_volumn']

        first_all_data = first_all_data.sort_index(ascending=False)
        second_all_data = second_all_data.sort_index(ascending=False)
        close_data = close_data.sort_index(ascending=False)

        wb = oxl.Workbook()
        ws = wb.create_sheet(index=0,title='oxl-sheet')

        columns = second_all_data.columns
        index = first_all_data.index
        row_number = len(index)
        col_number = len(columns)

        ws.cell(row=2,column=1).value = '日期'
        ws.cell(row=2,column=2).value = '收盘价'
        #  ws.cell(row=2,column=3).value = '股票代码'
        #  ws.cell(row=2,column=4).value = '股票名称'
        ws.cell(row=2,column=3).value = '中央结算系统持股量'
        ws.cell(row=2,column=4).value = '参与者数目'
        ws.cell(row=2,column=5).value = '总数百分比'
        ws.cell(row=2,column=6).value = '全部持股量'

        for i in range(row_number):
            ws.cell(row=i+3,column=1).value = index[i].date().isoformat()
            ws.cell(row=i+3,column=2).value = close_data.iat[i]
            ws.cell(row=i+3,column=3).value = hold_volumn.iat[i]
            ws.cell(row=i+3,column=4).value = people_number.iat[i]
            ws.cell(row=i+3,column=5).value = hold_precent.iat[i]
            ws.cell(row=i+3,column=6).value = all_volumn.iat[i]

        for i in range(col_number):
            ws.cell(row=2,column=i+7).value = columns[i]
            ws.cell(row=1,column=i+7).value = name_dict[columns[i]]

        for i in range(row_number):
            for j in range(col_number):
                ws.cell(row=3+i,column=j+7).value = second_all_data.iat[i,j]

        wb.save(sh_code+'.xlsx')
        #  contents.SetValue(r"数据已写入excel中!")
        contents.AppendText(sh_code+"数据已写入excel中!\n")

    contents.AppendText("全部数据已写入excel中!\n")

def get_close_data(config_dict):
    '''
    获取收盘价数据
    '''

    close_data = config_dict['close_data']
    close_data = pd.read_json(close_data)
    close_data.index = close_data['date']
    #  close_data = close_data.sort_index()
    return close_data['close']

def get_second_all_data(config_dict):
    '''
    获取第二部分数据
    '''
    second_all_data = config_dict['second_all_data']
    second_all_data = pd.read_json(second_all_data)
    #  second_all_data = second_all_data.sort_index()

    return second_all_data

def get_first_all_data():
    '''
    获取第一部分数据
    '''

    first_all_data = config_dict['first_all_data']
    first_all_data = pd.read_json(first_all_data)
    #  first_all_data = first_all_data.sort_index()
    return first_all_data


def plot_10(event):
    '''
    画前10大持仓股的曲线图
    '''

    symbol_number = len(sh_code_list)

    for j in range(symbol_number):
        sh_code = sh_code_list[j]
        config_dict = load_config_data(sh_code)

        close_data = get_close_data(config_dict)
        second_all_data = get_second_all_data(config_dict)
        name_dict = config_dict['name_dict']

        for i in range(len(second_all_data.columns)):
            if np.issubdtype(second_all_data.iloc[:,i],np.object_):
                second_all_data.iloc[:,i] = pd.to_numeric(second_all_data.iloc[:,i].str.replace(',', ''))

        close_data = close_data.sort_index()
        second_all_data = second_all_data.sort_index()
        second_all_data = second_all_data.sort_values(second_all_data.index[-1],axis=1,ascending=False)

        myfont = mpl.font_manager.FontProperties(fname="simhei.ttf")
        mpl.rcParams['axes.unicode_minus'] = False

        for i in range(10):
            temp_col = second_all_data.iloc[:,i]
            fig, ax1 = plt.subplots()
            #plt.plot(l2,lw=1.5, label='close')
            temp_col.plot(lw=1.5,label='volumn')
            plt.grid(True)
            plt.legend(loc=2)
            plt.axis('tight')
            plt.xlabel('time')
            plt.ylabel('volumn')

            ax2 = ax1.twinx()
            plt.plot(close_data, 'g',lw=1.5, label='close')
            plt.legend(loc=1)
            plt.ylabel('close')
            plt.title(name_dict[temp_col.name],fontproperties=myfont)
            plt.savefig(str(i)+'.png')

        contents.AppendText(sh_code + "画图成功!\n")

    contents.AppendText("画图完成!\n")

def ui_end_date_evt_text(event):
    '''
    起始输入日期的回调函数
    '''
    temp_string = ui_end_date.GetValue()

    if len(temp_string) == 8:
        temp_string = temp_string[0:4] + '-' + temp_string[4:6] + '-' + temp_string[6:8]
        temp_string = datetime.datetime.strptime(temp_string,"%Y-%m-%d").date()
        if temp_string > max_date:
            #  contents.SetValue("输入的日期小于最早可以得到的数据,网站最早可以得到1年以前的数据")
            contents.AppendText("输入的日期大于最晚可以得到的数据\n")
            temp_string = max_date
        ui_end_date.SetValue(temp_string.isoformat())

def ui_begin_date_evt_text(event):
    '''
    最终日期的回掉函数
    '''
    temp_string = ui_begin_date.GetValue()

    if len(temp_string) == 8:
        temp_string = temp_string[0:4] + '-' + temp_string[4:6] + '-' + temp_string[6:8]
        temp_string = datetime.datetime.strptime(temp_string,"%Y-%m-%d").date()
        if temp_string < min_date:
            #  contents.SetValue("输入的日期小于最早可以得到的数据,网站最早可以得到1年以前的数据")
            contents.AppendText("输入的日期小于最早可以得到的数据\n")
            temp_string = min_date
        ui_begin_date.SetValue(temp_string.isoformat())


def init():
    '''
    初始化数据变量
    '''
    contents.AppendText('日期格式的填写为:20180101\n')
    contents.AppendText('网站当前可以下载的时间区间为:%s至%s\n' % (min_date.isoformat(),max_date.isoformat()))
    ui_begin_date.SetValue(min_date.isoformat())
    ui_end_date.SetValue(max_date.isoformat())
    if os.path.exists('config.json'):
        with open('config.json','r',encoding='utf-8') as f:
            config_dict = json.load(f)
            temp_trade_cal = config_dict['trade_cal']
            temp_trade_cal = pd.read_json(temp_trade_cal)
            temp_trade_cal = temp_trade_cal.sort_index()

    else:
        temp_trade_cal = ts.trade_cal()
        config_dict = {}
        config_dict['trade_cal'] = temp_trade_cal.to_json()
        
        with open('config.json','w',encoding='utf-8') as f:
            json.dump(config_dict,f)

    return  temp_trade_cal

if __name__ == '__main__':

    url = r"http://www.hkexnews.hk/sdw/search/searchsdw_c.aspx"

    today_date = datetime.date.today()
    max_date = today_date - datetime.timedelta(1)
    min_date = datetime.date(today_date.year-1,today_date.month,today_date.day)

    app = wx.App()

    win = wx.Frame(None,title="simple editor",size=(830,800))

    begin_button = wx.Button(win,label='更新数据',pos=(160,60),size=(60,45))
    begin_button.Bind(wx.EVT_BUTTON,update)
    save_button = wx.Button(win,label='写入excle',pos=(240,60),size=(60,45))
    save_button.Bind(wx.EVT_BUTTON,write_to_excel)
    plot_button = wx.Button(win,label='画图',pos=(320,60),size=(60,45))
    plot_button.Bind(wx.EVT_BUTTON,plot_10)

    load_data_button = wx.Button(win,label='加载数据',pos=(160,100),size=(60,45))
    load_data_button.Bind(wx.EVT_BUTTON,load_symbol_data)

    ui_label1 = wx.StaticText(win, label = "起始日期", pos = (80,18))
    ui_label2 = wx.StaticText(win, label = "终止日期", pos = (280,18))

    ui_begin_date = wx.TextCtrl(win,pos=(140,15),size=(90,25))
    ui_begin_date.Bind(wx.EVT_TEXT,ui_begin_date_evt_text)
    ui_end_date = wx.TextCtrl(win,pos=(340,15),size=(90,25))
    ui_end_date.Bind(wx.EVT_TEXT,ui_end_date_evt_text)

    # 输入股票代码
    ui_label3 = wx.StaticText(win, label = "股票代码列表", pos = (100,140))
    symbol_list = wx.TextCtrl(win,pos=(100,170),size=(100,200),style=wx.TE_MULTILINE)

    ui_label4 = wx.StaticText(win, label = "操作记录", pos = (300,140))
    contents = wx.TextCtrl(win,pos=(300,170),size=(390,200),style=rt.RE_READONLY | wx.TE_MULTILINE)
    global_trade_cal = init()

    win.Show()
    app.MainLoop()
