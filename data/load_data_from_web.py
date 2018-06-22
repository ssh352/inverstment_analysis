#!/usr/bin/env python
# coding:utf-8

"""
function: 
@author: zkang kai
@contact: 474918208@qq.com
"""

import requests
import pandas as pd
from bs4 import BeautifulSoup
import tushare as ts
import openpyxl as oxl
import os
from collections import OrderedDict
import datetime
import json
import wx
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

url = r"http://www.hkexnews.hk/sdw/search/searchsdw_c.aspx"

def update(event):
    '''
    更新数据
    '''
    name_dict = config_dict.get('name_dict',{})
    first_all_data = config_dict['first_all_data']
    first_all_data = pd.read_json(first_all_data)
    second_all_data = config_dict['second_all_data']
    second_all_data = pd.read_json(second_all_data)

    begin_date = ui_begin_date.GetValue()
    begin_date = begin_date[0:4] + '-' + begin_date[4:6] + '-' + begin_date[6:8]
    end_date = ui_end_date.GetValue()
    end_date = end_date[0:4] + '-' + end_date[4:6] + '-' + end_date[6:8]

    already_begin_date = config_dict.get('already_begin_date',begin_date)
    already_end_date = config_dict.get('already_end_date',end_date)

    trade_cal = config_dict['trade_cal']
    trade_cal = pd.read_json(trade_cal)
    trade_cal = trade_cal[trade_cal['calendarDate'] >= begin_date]
    trade_cal = trade_cal[trade_cal['calendarDate'] <= end_date]
    trade_cal = trade_cal[trade_cal['isOpen'] == 1]
    trade_cal = trade_cal['calendarDate']

    browser = webdriver.Chrome()
    browser.get(url)

    for temp_time in trade_cal:

        if temp_time in config_dict:
            continue
    
        select = Select(browser.find_element_by_name('ddlShareholdingMonth'))
        select.select_by_value(temp_time[5:7])

        select = Select(browser.find_element_by_name('ddlShareholdingDay'))
        select.select_by_value(temp_time[8:10])

        select = Select(browser.find_element_by_name('ddlShareholdingYear'))
        select.select_by_value(temp_time[0:4])

        element = browser.find_element_by_id("txtStockCode")
        element.send_keys("91888")

        browser.find_element_by_id("btnSearch").click()

        html_txt = browser.page_source

        soup = BeautifulSoup(html_txt,'html.parser')

        first_part = get_first_part(soup,temp_time)
        second_part = get_second_part(soup)
        second_part.index = [temp_time]
        first_all_data = first_part.append(first_all_data) 
        second_all_data = second_part.append(second_all_data) 

        config_dict[temp_time] = [first_part.to_json(),second_part.to_json()] 
        browser.back() 

    
    first_all_data = first_all_data.sort_index(ascending=False)
    second_all_data = second_all_data.sort_index(ascending=False)

    already_begin_date = first_all_data.index[-1]
    already_end_date = first_all_data.index[0]
    close_data = ts.get_k_data('601888',start=already_begin_date,end=already_end_date)

    config_dict['name_dict'] = name_dict
    config_dict['close_data'] = close_data.to_json()
    config_dict['already_begin_date'] = already_begin_date 
    config_dict['already_end_date'] = already_end_date
    config_dict['first_all_data'] = first_all_data.to_json()
    config_dict['second_all_data'] = second_all_data.to_json()

    with open('config.json','w',encoding='utf-8') as f:
        json.dump(config_dict,f)

    print('数据从%s至%s!' % (already_begin_date,already_end_date))

def load_from_local():
    
    with open('__ HKEX __ HKEXnews __.html','r',encoding="utf-8") as f:
        return f.read()

def get_first_part(soup,temp_time):

    temp_html = soup.find("table",{"id":"Table5"})
    temp_html = temp_html.find_all('td',{"class":"arial12black"})

    first_dict = {}
    temp_string = temp_html[1].get_text().strip()
    temp_string = datetime.datetime.strptime(temp_string,"%d/%m/%Y")
    temp_string = temp_string.strftime("%Y-%m-%d")
    #  first_dict['last_date'] = temp_string
    first_dict['stock_code'] = temp_html[3].get_text().strip()
    first_dict['stock_name'] = temp_html[5].get_text().strip()

    temp_html = soup.find("div",{"id":{"pnlResultSummary"}})
    temp_html = temp_html.find_all("span",{"class":"mobilezoom"})

    first_dict = {}
    first_dict['hold_volumn'] = temp_html[0].get_text().strip()
    first_dict['people_number'] = temp_html[1].get_text().strip()
    first_dict['hold_precent'] = temp_html[2].get_text().strip()
    first_dict['all_volumn'] = temp_html[6].get_text().strip()

    return pd.DataFrame(first_dict,index=[temp_time])

def get_second_part(soup):

    name_dict = config_dict.get('name_dict',{})
    temp_html = soup.find("table",{"id":{"participantShareholdingList"}})
    temp_html = temp_html.find_all('tr')

    row_number = len(temp_html)
    col_number = len(temp_html[0].find_all('td'))

    d = {}
    for i in range(col_number):
        d[i] = []

    for i in range(1,row_number):
        temp_tr = temp_html[i]
        temp_td = temp_tr.find_all('td')
    
        if len(temp_td) !=col_number:
            continue
        
        for j in range(col_number):
            temp_string = temp_td[j].get_text().strip()
            d[j].append(temp_string)

    last_pd = pd.DataFrame(d)
    
    last_dict = {}
    for i in range(len(last_pd)):
        last_dict[last_pd.iat[i,0]] = last_pd.iat[i,1]

    name_dict.update(last_dict)

    new_pd = pd.DataFrame(last_pd[3])
    new_pd = new_pd.T
    new_pd.columns = last_pd[0]

    return new_pd

def write_to_excel(event):
    '''
    把数据写入excle中
    '''
    name_dict = config_dict.get('name_dict',{})
    first_all_data = config_dict['first_all_data']
    first_all_data = pd.read_json(first_all_data)
    second_all_data = config_dict['second_all_data']
    second_all_data = pd.read_json(second_all_data)
    close_data = config_dict['close_data']
    close_data = pd.read_json(close_data)
    close_data = close_data['close']
    hold_volumn = first_all_data['hold_volumn']
    people_number = first_all_data['people_number']
    hold_precent = first_all_data['hold_precent']
    all_volumn = first_all_data['all_volumn']

    wb = oxl.Workbook()
    ws = wb.create_sheet(index=0,title='oxl-sheet')

    columns = second_all_data.columns
    index = first_all_data.index
    row_number = len(index)
    col_number = len(columns)
    
    ws.cell(row=2,column=1).value = '日期'
    ws.cell(row=2,column=2).value = '收盘价'
    #  ws.cell(row=2,column=3).value = '股票代码'
    #  ws.cell(row=2,column=4).value = '股票名称'
    ws.cell(row=2,column=3).value = '中央结算系统持股量'
    ws.cell(row=2,column=4).value = '参与者数目'
    ws.cell(row=2,column=5).value = '总数百分比'
    ws.cell(row=2,column=6).value = '全部持股量'
    
    for i in range(row_number):
        ws.cell(row=i+3,column=1).value = str(index[i])[0:10]
        ws.cell(row=i+3,column=2).value = close_data.iat[i]
        ws.cell(row=i+3,column=3).value = hold_volumn.iat[i]
        ws.cell(row=i+3,column=4).value = people_number.iat[i]
        ws.cell(row=i+3,column=5).value = hold_precent.iat[i]
        ws.cell(row=i+3,column=6).value = all_volumn.iat[i]

    for i in range(col_number):
        ws.cell(row=2,column=i+7).value = columns[i]
        ws.cell(row=1,column=i+7).value = name_dict[columns[i]] 

    for i in range(row_number):
        for j in range(col_number):
            ws.cell(row=3+i,column=j+7).value = second_all_data.iat[i,j]

    wb.save('test.xlsx')
    print("数据已写入excel中!")

def plot_10(merge_pd):
    '''
    画前10大持仓股的曲线图
    '''
    temp_pd = merge_pd.sort_values(merge_pd.index[0],axis=1,ascendind=False)
    
    number = min(10,len(merge_pd))
    
    for i in range(number):
        pass

def save_to_config():

    first_dict = get_first_part()
    last_date = first_dict['last_date']

    if str(pd.Timestamp(last_date)) in close_data:
        print("该日期数据已经存在!")
        save_to_excel(config_dict['first_dict'],config_dict['second_dict'],pd.read_json(config_dict['old_pd'])) 
        return

    second_dict = get_second_part()
    second_pd = get_second_part(first_dict)
    merge_pd = second_pd.append(old_pd)
    merge_pd = merge_pd.sort_index(ascending=False)

    config_dict['name_dict'] = name_dict
    config_dict['close_data'] = close_data
    config_dict[last_date] = [first_dict,second_dict,second_pd.to_json()]
    config_dict['old_pd'] = merge_pd.to_json()
    config_dict['first_dict'] = first_dict
    config_dict['second_dict'] = second_dict
    config_dict['second_pd'] = second_pd.to_json()

    save_to_excel(first_dict,second_dict,merge_pd)
    
    with open('config.json','w',encoding='utf-8') as f:
        json.dump(config_dict,f)

    print('数据更新至%s!' % merge_pd.index[0])


def init():
    '''
    初始化数据变量
    '''
    if os.path.exists('config.json'):
        with open('config.json','r',encoding='utf-8') as f:
            config_dict = json.load(f)
            print("当前拥有数据区间为%s至%s" % (config_dict['already_begin_date'],config_dict['already_end_date']))
    else:
        config_dict = {}
        config_dict['trade_cal'] = ts.trade_cal().to_json()
        config_dict['name_dict'] = {}
        first_all_data = pd.DataFrame()
        config_dict['first_all_data'] = first_all_data.to_json()
        second_all_data = pd.DataFrame()
        config_dict['second_all_data'] = second_all_data.to_json()
        print("当前未拥有任何数据")

    return config_dict

if __name__ == '__main__':

    config_dict = init()

    app = wx.App()

    win = wx.Frame(None,title="simple editor",size=(410,335))

    begin_button = wx.Button(win,label='更新数据',pos=(225,10),size=(80,45))
    begin_button.Bind(wx.EVT_BUTTON,update)
    save_button = wx.Button(win,label='写入excle',pos=(315,10),size=(80,45))
    save_button.Bind(wx.EVT_BUTTON,write_to_excel)

    ui_label1 = wx.StaticText(win, label = "起始日期", pos = (5,5)) 
    ui_label2 = wx.StaticText(win, label = "终止日期", pos = (5,35)) 

    ui_begin_date = wx.TextCtrl(win,pos=(60,5),size=(150,25))
    ui_end_date = wx.TextCtrl(win,pos=(60,35),size=(150,25))

    contents = wx.TextCtrl(win,pos=(5,70),size=(390,260),style=wx.TE_MULTILINE |
                           wx.HSCROLL)

    win.Show()
    app.MainLoop()
